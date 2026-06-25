import os, canopen
from ros2_canbus.JS2_Motor_CANOpen_Lib_V_1_0.Housekeeping.Logger_Lib import Logger
# from CANopen_Network.Network_Lib import CANopen_Network

class Load_Settings():
    def __init__(self, settings_file: str,
                Node_Name: str):
                # canopen_handle: CANopen_Network):
        try:
            self.settings_file = settings_file
            self.Node_Name = Node_Name
            # self.canopen_handle = canopen_handle
            self.Node_ID = -1
            self.log = Logger(self.Node_Name, self.Node_ID)
            self.BASE_DIR = os.path.dirname(os.path.abspath(__file__))
            self.motor_excel_file = os.path.abspath(os.path.join("/home/jontro_soinik_2_0-2/ros2_ws/src/ros2_canbus/ros2_canbus/JS2_Motor_CANOpen_Lib_V_1_0/Motor_Settings", settings_file))
            
            # print(self.motor_excel_file)
            self.settings = self.load_file(self.motor_excel_file)
            motor_specific_settings = self.find_motor_settings(self.Node_Name)
            self.load_motor_specific_settings(motor_specific_settings)
            if self.Node_ID <= 0:
                raise ValueError("Motor Name not found or invalid ID for motor")
            else:
                self.log = Logger(self.Node_Name, self.Node_ID)
                self.log.print(f"Settings loaded","__init__")

        except Exception as e:
            self.log.print(f"Error reading motor config for '{self.Node_Name}': {e}","❌", "Load_Settings","__init__", f"{e}")



    
    def load_file(self, settings_file):
        """Extended to load all settings from the Excel file."""
        import os
        from openpyxl import load_workbook

        ext = os.path.splitext(settings_file)[1].lower()
        if ext != ".xlsx":
            raise RuntimeError(f"Only .xlsx is supported for now. Got: {settings_file}")

        wb = load_workbook(settings_file, data_only=True)
        
        if "nodes" not in wb.sheetnames:
            raise RuntimeError("Excel must contain a sheet named 'nodes'.")

        ws = wb["nodes"]
        
        header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if not header_cells:
            raise RuntimeError("First row must contain headers: node_name, node_id, eds_file, etc.")

        headers = [str(h).strip().lower() if h is not None else "" for h in header_cells]
        col = {h: i for i, h in enumerate(headers)}

        # Columns for the new settings
        """ PLEASE USE AI TO EDIT THE FOLLOWING CODE. IT'S VERY LENGTHY AND BORING"""
        H_NAME = col.get("node_name")
        H_ID = col.get("node_id")
        H_EDS = col.get("eds_file")
        H_BLD_MAX_VEL = col.get("bldc_max_velocity")
        H_ENCODER_BIT = col.get("encoder_bit")
        H_GEAR_RATIO = col.get("gear_ratio")
        H_TIME_PERIOD_MS = col.get("time_period_ms")
        H_BLD_VELOCITY_UNIT = col.get("bldc_velocity_unit")
        H_BLD_ACC_UNIT = col.get("bldc_acceleration_unit")
        H_BLD_DEC_UNIT = col.get("bldc_deceleration_unit")
        H_TORQUE_CURR_RATIO = col.get("torque_current_ratio")
        H_MAX_POS = col.get("max_position")
        H_MIN_POS = col.get("min_position")
        H_MAX_VEL = col.get("max_velocity")
        H_MIN_VEL = col.get("min_velocity")
        H_MAX_ACC = col.get("max_acceleration")
        H_MIN_ACC = col.get("min_acceleration")
        H_CURR_ACC = col.get("current_acceleration")
        H_MAX_DEC = col.get("max_deceleration")
        H_MIN_DEC = col.get("min_deceleration")
        H_CURR_DEC = col.get("current_deceleration")
        H_MAX_TRQUE = col.get("max_torque")
        H_MIN_TORQUE = col.get("min_torque")
        H_TORQUE_SLOPE_MAX = col.get("torque_slope_max")
        H_TORQUE_SLOPE_MIN = col.get("torque_slope_min")
        H_CURR_TORQUE_SLOPE = col.get("current_torque_slope")
        H_MODE = col.get("mode")
        H_FEEDBACK = col.get("feedback")
        H_HEARTBEAT_TIMOUT_MS = col.get("heartbeat_timeout_ms")
        H_DEBUG_INIT = col.get("debug_init")
        H_DEBUG_COMMAND = col.get("debug_command")
        H_DEBUG_FEEDBACK = col.get("debug_feedback")
        H_DEBUG_HEARTBEAT = col.get("debug_heartbeat")
        H_DEBUG_CONTROLWORD = col.get("debug_controlword")
        H_Failure_Exit = col.get("failure_exit")

        # Validate required fields
        if H_NAME is None:
            raise RuntimeError("Missing required header 'node_name'.")
        
        nodes = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue

            nm = r[H_NAME] if H_NAME < len(r) else None
            name = str(nm).strip() if nm else ""
            if not name:
                continue  # skip blank-name rows

            node_id = r[H_ID] if H_ID is not None else None
            eds_file = r[H_EDS] if H_EDS is not None else None
            bldc_max_velocity = r[H_BLD_MAX_VEL] if H_BLD_MAX_VEL is not None else None
            encoder_bit = r[H_ENCODER_BIT] if H_ENCODER_BIT is not None else None
            gear_ratio = r[H_GEAR_RATIO] if H_GEAR_RATIO is not None else None
            time_period_ms = r[H_TIME_PERIOD_MS] if H_TIME_PERIOD_MS is not None else None
            bldc_velocity_unit = r[H_BLD_VELOCITY_UNIT] if H_BLD_VELOCITY_UNIT is not None else None
            bldc_acceleration_unit = r[H_BLD_ACC_UNIT] if H_BLD_ACC_UNIT is not None else None
            bldc_deceleration_unit = r[H_BLD_DEC_UNIT] if H_BLD_DEC_UNIT is not None else None
            torque_current_ratio = r[H_TORQUE_CURR_RATIO] if H_TORQUE_CURR_RATIO is not None else None
            max_position = r[H_MAX_POS] if H_MAX_POS is not None else None
            min_position = r[H_MIN_POS] if H_MIN_POS is not None else None
            max_velocity = r[H_MAX_VEL] if H_MAX_VEL is not None else None
            min_velocity = r[H_MIN_VEL] if H_MIN_VEL is not None else None
            max_acceleration = r[H_MAX_ACC] if H_MAX_ACC is not None else None
            min_acceleration = r[H_MIN_ACC] if H_MIN_ACC is not None else None
            current_acceleration = r[H_CURR_ACC] if H_CURR_ACC is not None else None
            max_deceleration = r[H_MAX_DEC] if H_MAX_DEC is not None else None
            min_deceleration = r[H_MIN_DEC] if H_MIN_DEC is not None else None
            current_deceleration = r[H_CURR_DEC] if H_CURR_DEC is not None else None
            max_torque = r[H_MAX_TRQUE] if H_MAX_TRQUE is not None else None
            min_torque = r[H_MIN_TORQUE] if H_MIN_TORQUE is not None else None

            
            torque_slope_max = r[H_TORQUE_SLOPE_MAX] if H_TORQUE_SLOPE_MAX is not None else None
            torque_slope_min = r[H_TORQUE_SLOPE_MIN] if H_TORQUE_SLOPE_MIN is not None else None
            current_torque_slope = r[H_CURR_TORQUE_SLOPE] if H_CURR_TORQUE_SLOPE is not None else None
            mode = r[H_MODE] if H_MODE is not None else None
            feedback = bool(r[H_FEEDBACK]) if H_FEEDBACK is not None else True
            heartbeat_timeout_ms = r[H_HEARTBEAT_TIMOUT_MS] if H_HEARTBEAT_TIMOUT_MS is not None else None
            debug_init = bool(r[H_DEBUG_INIT]) if H_DEBUG_INIT is not None else True
            debug_command = bool(r[H_DEBUG_COMMAND]) if H_DEBUG_COMMAND is not None else True
            debug_feedback = bool(r[H_DEBUG_FEEDBACK]) if H_DEBUG_FEEDBACK is not None else True
            debug_heartbeat = bool(r[H_DEBUG_HEARTBEAT]) if H_DEBUG_HEARTBEAT is not None else True
            debug_controlword = bool(r[H_DEBUG_CONTROLWORD]) if H_DEBUG_CONTROLWORD is not None else True
            failure_exit = bool(r[H_Failure_Exit]) if H_Failure_Exit is not None else False

            # Collect node data
            nodes.append({
                "node_name": name,
                "node_id": node_id,
                "eds_file": eds_file,
                "BLDC_max_velocity": bldc_max_velocity,
                "encoder_bit": encoder_bit,
                "gear_ratio": gear_ratio,
                "time_period_ms": time_period_ms,
                "BLDC_velocity_unit": bldc_velocity_unit,
                "BLDC_acceleration_unit": bldc_acceleration_unit,
                "BLDC_deceleration_unit": bldc_deceleration_unit,
                "torque_current_ratio": torque_current_ratio,
                "max_position": max_position,
                "min_position": min_position,
                "max_velocity": max_velocity,
                "min_velocity": min_velocity,
                "max_acceleration": max_acceleration,
                "min_acceleration": min_acceleration,
                "current_acceleration": current_acceleration,
                "max_deceleration": max_deceleration,
                "min_deceleration": min_deceleration,
                "current_deceleration": current_deceleration,
                "max_torque":max_torque,
                "min_torque": min_torque,
                "torque_slope_max": torque_slope_max,
                "torque_slope_min": torque_slope_min,
                "current_torque_slope": current_torque_slope,
                "mode": mode,
                "feedback": feedback,
                "heartbeat_timeout_ms": heartbeat_timeout_ms,
                "DEBUG_INIT": debug_init,
                "DEBUG_COMMAND": debug_command,
                "DEBUG_FEEDBACK": debug_feedback,
                "DEBUG_HEARTBEAT": debug_heartbeat,
                "DEBUG_CONTROLWORD": debug_controlword,
                "FAILURE_EXIT": failure_exit
            })
            
        return {"nodes": nodes}
    
    def find_motor_settings(self, _node_name):
        try:
            nodes = self.settings.get('nodes', [])
            for node in nodes:
                if node.get('node_name') == _node_name:
                    self.Node_ID  = node.get('node_id', -1)
                    self.eds_file = (node.get('eds_file') or "").strip()
                    if self.Node_ID is None or not self.eds_file:
                        raise RuntimeError(
                            f"Settings for '{_node_name}' missing required fields: "
                            f"node_id={self.Node_ID}, eds_file='{self.eds_file}'"
                        )
                    return node
        except Exception as e:
            self.log.print(f"Error reading motor config for '{_node_name}': {e}","❌","Load_Settings","find_motor_settings" f"{e}")
            raise
        raise ValueError(f"Node '{_node_name}' not found in settings file.")

    def load_motor_specific_settings(self, settings):
        try:
            self.mode = settings.get('mode')
            self.feedback =  settings.get("feedback")
            self.heartbeat_timeout_ms = settings.get("heartbeat_timeout_ms")

            self.DEBUG_INIT =  settings.get("DEBUG_INIT")
            self.DEBUG_COMMAND =  settings.get("DEBUG_COMMAND")
            self.DEBUG_FEEDBACK =  settings.get("DEBUG_FEEDBACK")
            self.DEBUG_HEARTBEAT =  settings.get("DEBUG_HEARTBEAT")
            self.DEBUG_CONTROLWORD =  settings.get("DEBUG_CONTROLWORD")
            self.FAILURE_EXIT =  settings.get("FAILURE_EXIT")

            # Initialize motion parameters from Excel or use default values
            self.BLDC_max_velocity = settings.get('BLDC_max_velocity')
            # print(self.BLDC_max_velocity)
            self.encoder_bit = settings.get('encoder_bit')
            # print(self.encoder_bit)
            self.gear_ratio = settings.get('gear_ratio')
            # print(self.gear_ratio)
            self.time_period_ms = settings.get('time_period_ms')
            # print(self.time_period_ms)

            self.BLDC_velocity_unit = settings.get('BLDC_velocity_unit')
            # print(self.BLDC_velocity_unit)
            self.BLDC_acceleration_unit = settings.get('BLDC_acceleration_unit')
            # print(self.BLDC_acceleration_unit)
            self.BLDC_deceleration_unit = settings.get('BLDC_deceleration_unit')
            # print(self.BLDC_deceleration_unit)
            self.torque_current_ratio = settings.get('torque_current_ratio')
            # print(self.torque_current_ratio)

            # Position, velocity, and acceleration limits from Excel
            self.max_position = settings.get('max_position')
            # print(self.max_position)
            self.min_position = settings.get('min_position')
            # print(self.min_position)
            self.max_velocity = settings.get('max_velocity')
            # print(self.max_velocity)
            self.min_velocity = settings.get('min_velocity')
            # print(self.min_velocity)
            self.max_acceleration = settings.get('max_acceleration')
            # print(self.max_acceleration)
            self.min_acceleration = settings.get('min_acceleration')
            # print(self.min_acceleration)
            self.current_acceleration = settings.get('current_acceleration')
            # print(self.current_acceleration)
            self.max_deceleration = settings.get('max_deceleration')
            # print(self.max_deceleration)
            self.min_deceleration = settings.get('min_deceleration')
            # print(self.min_deceleration)
            self.current_deceleration = settings.get('current_deceleration')
            # print(self.current_deceleration)
            self.max_torque = settings.get('max_torque')
            # print(self.max_torque)
            self.min_torque = settings.get('min_torque')
            # print(self.min_torque)
            self.torque_slope_max = settings.get('torque_slope_max')
            # print(self.torque_slope_max)
            self.torque_slope_min = settings.get('torque_slope_min')
            # print(self.torque_slope_min)
            self.current_torque_slope = settings.get('current_torque_slope')

        except Exception as e:
            self.log.print(f"Error reading motor config for '{self.Node_Name}': {e}","❌" "Load_Settings","load_motor_specific_settings", f"{e}")

            pass

