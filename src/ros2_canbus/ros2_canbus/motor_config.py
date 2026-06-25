"""
motor_config.py
===============
MotorConfig data class (loaded from motor_settings.xlsx) and the
``load_motor_configs()`` loader.

Conversion methods on MotorConfig delegate to :mod:`conversions`.
"""

import openpyxl
from ros2_canbus.conversions import (
    rad_to_counts, counts_to_rad,
    rad_per_s_to_can_vel, can_vel_to_rad_per_s,
    clamp,
)


# ── CiA-402 modes ──────────────────────────────────────────────────────

MODE_CSP = 8
MODE_CSV = 9
MODE_CST = 10


# ── MotorConfig ────────────────────────────────────────────────────────

class MotorConfig:
    """Per-motor parameters parsed from one row of ``motor_settings.xlsx``."""

    __slots__ = (
        'name', 'node_id', 'bus', 'mode',
        'encoder_bit', 'gear_ratio', 'velocity_unit',
        'counts_per_output_rev',
        'max_position', 'min_position',
        'max_velocity', 'min_velocity',
        'max_torque',   'min_torque',
    )

    def __init__(self, row: dict):
        self.name          = row['node_name']
        self.node_id       = int(row['node_id'])
        self.bus            = row['bus']
        self.mode           = int(row['mode'])
        self.encoder_bit    = int(row['encoder_bit'])
        self.gear_ratio     = float(row['gear_ratio'])
        self.velocity_unit  = float(row['BLDC_velocity_unit'])
        self.counts_per_output_rev = (1 << self.encoder_bit) * self.gear_ratio

        self.max_position = float(row['max_position'])
        self.min_position = float(row['min_position'])
        self.max_velocity = float(row['max_velocity'])
        self.min_velocity = float(row['min_velocity'])
        self.max_torque   = float(row['max_torque'])
        self.min_torque   = float(row['min_torque'])

    # -- conversion shortcuts (delegate to conversions.py) ---------------

    def rad_to_counts(self, q_rad: float) -> int:
        return rad_to_counts(q_rad, self.counts_per_output_rev)

    def counts_to_rad(self, cnt: float) -> float:
        return counts_to_rad(cnt, self.counts_per_output_rev)

    def rad_per_s_to_can_vel(self, w_rad_s: float) -> int:
        return rad_per_s_to_can_vel(w_rad_s, self.gear_ratio)

    def can_vel_to_rad_per_s(self, v_can: float) -> float:
        return can_vel_to_rad_per_s(v_can, self.gear_ratio)

    def clamp_position_cnt(self, cnt: int) -> int:
        return clamp(cnt, self.min_position, self.max_position)

    def clamp_velocity_cnt(self, v: int) -> int:
        return clamp(v, self.min_velocity, self.max_velocity)

    def clamp_torque_cnt(self, t: int) -> int:
        return clamp(t, self.min_torque, self.max_torque)


# ── xlsx loader ────────────────────────────────────────────────────────

def load_motor_configs(xlsx_path: str, bus_filter: set) -> dict:
    """Return ``{motor_name: MotorConfig}`` for every motor whose
    ``node_name`` is in *bus_filter*."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'nodes' not in wb.sheetnames:
        raise RuntimeError(
            f"motor_settings.xlsx missing 'nodes' sheet; found: {wb.sheetnames}")
    ws = wb['nodes']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    cfgs = {}
    for row in rows[1:]:
        rd = dict(zip(header, row))
        name = rd.get('node_name')
        if not name or rd.get('bus') in (None, ''):
            continue
        if name not in bus_filter:
            continue
        cfgs[name] = MotorConfig(rd)
    return cfgs
